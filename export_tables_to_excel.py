from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path

import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, text


WORKSPACE_ROOT = Path(__file__).resolve().parent
LOCAL_PYTHON_PACKAGES = WORKSPACE_ROOT / ".python_packages"

if LOCAL_PYTHON_PACKAGES.exists():
    import sys

    sys.path.insert(0, str(LOCAL_PYTHON_PACKAGES))


def _sheet_name(table_name: str) -> str:
    return table_name[:31]


TABLE_DESCRIPTIONS = {
    "article": "Product/article master listing for the manufactured device families.",
    "bom": "Bill of materials header tying an article and configuration to a BOM version.",
    "bom_node": "Hierarchy of assemblies and components inside each BOM.",
    "configuration": "Configuration variants for each article, such as market or option set.",
    "defect": "In-factory quality issues detected during production or testing.",
    "factory": "Factory master data.",
    "field_claim": "Customer-reported failures after shipment.",
    "line": "Production lines within a factory.",
    "part": "Individual physical part instances tied to a supplier batch and part master.",
    "part_master": "Part master catalog with canonical part numbers and descriptions.",
    "product": "Built product instances; central entity referenced by quality events.",
    "product_action": "Team-editable action log for investigations, initiatives, and 8D workflow.",
    "product_part_install": "Which physical parts were installed into which products and positions.",
    "production_order": "Production orders used to build products.",
    "rework": "Corrective actions performed to address defects.",
    "section": "Manufacturing/test sections within a production line.",
    "supplier_batch": "Supplier lot or batch records for incoming parts.",
    "test": "Test definitions, limits, and target locations or parts.",
    "test_result": "Measured test outcomes for a specific product and test run.",
}


def _format_datetime_columns(df: pd.DataFrame) -> pd.DataFrame:
    formatted = df.copy()
    for column in formatted.columns:
        series = formatted[column]
        if pd.api.types.is_datetime64_any_dtype(series):
            formatted[column] = series.dt.strftime("%Y-%m-%d %H:%M:%S").where(
                series.notna(), None
            )
    return formatted


def _autosize_worksheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)


def _style_worksheet(worksheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    _autosize_worksheet(worksheet)


def _add_summary_sheet(writer: pd.ExcelWriter, table_frames: dict[str, pd.DataFrame]) -> None:
    summary_rows = []
    for table_name, df in table_frames.items():
        summary_rows.append(
            {
                "table_name": table_name,
                "row_count": len(df),
                "column_count": len(df.columns),
                "columns": ", ".join(df.columns),
                "description": TABLE_DESCRIPTIONS.get(table_name, ""),
            }
        )

    summary_df = pd.DataFrame(summary_rows)
    summary_df.to_excel(writer, sheet_name="Summary", index=False)
    worksheet = writer.book["Summary"]
    _style_worksheet(worksheet)

    for row_idx in range(2, len(summary_df) + 2):
        cell = worksheet.cell(row=row_idx, column=1)
        cell.hyperlink = f"#'{_sheet_name(cell.value)}'!A1"
        cell.style = "Hyperlink"


def main() -> None:
    load_dotenv(WORKSPACE_ROOT / ".env")
    database_url = os.getenv("MANEX_PG_URL")
    if not database_url:
        raise RuntimeError("Missing MANEX_PG_URL in .env")

    output_dir = WORKSPACE_ROOT / "outputs"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"manex_tables_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    engine = create_engine(database_url)

    with engine.connect() as conn:
        table_rows = conn.execute(
            text(
                """
                SELECT table_name
                FROM information_schema.tables
                WHERE table_schema = 'public'
                  AND table_type = 'BASE TABLE'
                ORDER BY table_name
                """
            )
        ).fetchall()
        table_names = [row.table_name for row in table_rows]
        table_frames = {
            table_name: pd.read_sql(text(f'SELECT * FROM "{table_name}"'), conn)
            for table_name in table_names
        }

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            _add_summary_sheet(writer, table_frames)

            for table_name in table_names:
                df = table_frames[table_name]
                export_df = _format_datetime_columns(df)
                export_df.to_excel(writer, sheet_name=_sheet_name(table_name), index=False)
                _style_worksheet(writer.book[_sheet_name(table_name)])

    print(output_path)


if __name__ == "__main__":
    main()
